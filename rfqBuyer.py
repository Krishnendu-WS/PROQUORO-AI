"""
RFQ Automation Script
=====================
- Auto-login (credentials from config or prompt once)
- Reads all RFQ data from RFQ_publish.csv
- Runs form-fill + navigation for each row (N iterations)
- Captures screenshot after each form is filled
- Generates Excel test report at the end
"""

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from datetime import datetime, timedelta
import re
import csv
import os
import sys
import time
import base64
from io import BytesIO

# ─── CONFIG ──────────────────────────────────────────────────────────────────
LOGIN_URL   = "https://stg.proquro.ai/sign-in?redirect=%2Fcompany-admin%2Frole-management"
LOGIN_EMAIL = "krish.ws@webspiders.com"   # ← set your email here
LOGIN_PASS  = "abcd@1234567"            # ← set your password here

CSV_FILE        = "RFQ_publish.csv"      # must be in same folder as this script
SCREENSHOT_DIR  = "rfq_screenshots"
REPORT_FILE     = "RFQ_Test_Report.xlsx"

# Fixed values that don't come from CSV
QUOTE_VALIDITY_DAYS = "15"
DELIVERY_LOCATION_OPTION = "TATA STEEL LIMITED-Registered"

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def load_csv(path: str) -> list[dict]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def take_screenshot(page, name: str) -> str:
    """Save a screenshot; return the file path."""
    os.makedirs(SCREENSHOT_DIR, exist_ok=True)
    safe = re.sub(r"[^\w\-]", "_", name)
    path = os.path.join(SCREENSHOT_DIR, f"{safe}.png")
    page.screenshot(path=path, full_page=True)
    return path


def img_to_base64(path: str) -> str | None:
    if not path or not os.path.exists(path):
        return None
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()


def set_datetime_picker(page, field_name: str, day: int, hour: str, minute: str, ampm: str):
    """Click a datetime button and fill in the picker."""
    page.get_by_role("button", name=field_name).click()
    page.get_by_role("button", name=str(day), exact=True).click()
    page.get_by_role("button", name="Hour").click()
    page.locator("div").filter(has_text=re.compile(f"^{hour}$")).click()
    page.get_by_role("button", name="Minute").click()
    page.locator("div").filter(has_text=re.compile(f"^{minute}$")).click()
    page.get_by_role("button", name="AM or PM").click()
    page.locator("div").filter(has_text=re.compile(f"^{ampm}$")).click()
    page.get_by_role("button", name="Apply").click()

def select_expected_delivery_date(page, days_after):
    """
    Select expected delivery date in Playwright.
    Automatically handles next month navigation.
    """
    # Calculate target date
    target_date = datetime.now() + timedelta(days=days_after)

    target_day = str(target_date.day)
    current_month = datetime.now().month
    target_month = target_date.month

    # Open date picker
    page.get_by_role("button", name="Expected Delivery Date").click()

    # If target date is in next month, click next month
    if target_month != current_month:
        page.get_by_role("button", name="Next month").click()

    # Select the day
    page.get_by_role(
        "button",
        name=target_day,
        exact=True
    ).click()


# ─── LOGIN ────────────────────────────────────────────────────────────────────

def login(page):
    page.goto(LOGIN_URL)
    page.get_by_role("button", name="Login").click()
    page.get_by_text("Securely Login With Email").click()

    # Fill credentials automatically

    page.get_by_role("textbox", name="Email").fill(LOGIN_EMAIL)
    page.get_by_role("textbox", name="Password").fill(LOGIN_PASS)
    
    page.get_by_role("button", name="Log In").click()

    # Wait for dashboard to load
    page.wait_for_load_state("networkidle", timeout=30_000)
    print("✅  Logged in successfully.")


# ─── SINGLE RFQ SUBMISSION ───────────────────────────────────────────────────

def submit_rfq(page, row: dict, row_index: int) -> dict:
    """
    Fill and submit one RFQ from a CSV row.
    Returns a result dict for the report.
    """
    result = {
        "Row": row_index,
        "RFQ Title": row.get("RFQ Title", ""),
        "RFx Type": row.get("RFx Type", ""),
        "Description": row.get("Description", ""),
        "Procurement Type": row.get("Procurement Type", ""),
        "Delivery Terms": row.get("Delivery Terms", ""),
        "Department": row.get("Department", ""),
        "Cost Center": row.get("Cost Center", ""),
        "Delivery Location": row.get("Delivery Location", ""),
        "Item Name": row.get("Item Name", ""),
        "Qty": row.get("Qty", ""),
        "Payment Terms": row.get("Payment Terms", ""),
        "Vendor": row.get("vendor", ""),
        "Status": "PASS",
        "Error": "",
        "Screenshot": "",
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    try:
        # ── Navigate to Create New RFQ ──────────────────────────────────────
        
        page.get_by_role("link", name="+ Create New RFQ").click()

        # ── Page 1: Header Info ─────────────────────────────────────────────

        # RFx Type
        page.get_by_role("button", name="Select RFx Type").click()
        page.get_by_text(row["RFx Type"]).click()

        # RFQ Title
        page.get_by_role("textbox", name=re.compile("Enter RFQ title", re.I)).fill(row["RFQ Title"])

        # Description
        page.get_by_role("textbox", name="Enter description").fill(row["Description"])

        # Procurement Type
        page.get_by_role("button", name=row["Procurement Type"]).click()     # open dropdown

        # Delivery Terms
        page.get_by_role("button", name="Select Delivery Terms").click()
        page.get_by_role("option", name=row["Delivery Terms"]).click()

        # Department
        page.get_by_role("button", name="Search and select department").click()
        page.get_by_role("textbox", name="Search departments...").fill(row["Department"])
        page.get_by_role("option", name=row["Department"]).click()

        # Cost Center
        page.get_by_role("button", name="Search and select cost center").click()
        page.get_by_role("textbox", name="Search cost centers...").fill(row["Cost Center"])
        page.get_by_role("option", name=row["Cost Center"]).click()

        # Calculate time values
        now = datetime.now()
        day, hour, minute, ampm = (
            now.day,
            now.strftime("%I"),
            now.strftime("%M"),
            now.strftime("%p"),
        )
        minute_plus_one = f"{(int(minute) + 1) % 60:02d}"
        minute_plus_six = f"{(int(minute) + 6) % 60:02d}"
        next_day        = (now + timedelta(days=1)).day
        

        # Publish Date
        set_datetime_picker(page, "Publish Date", day, hour, minute_plus_one, ampm)

        # Bid Start Date
        set_datetime_picker(page, "Bid Start Date", day, hour, minute_plus_six, ampm)

        # Bid End Date
        set_datetime_picker(page, "Bid End Date", next_day, hour, minute_plus_one, ampm)

        # Quote Validity
        page.get_by_placeholder("Enter number of days").fill(QUOTE_VALIDITY_DAYS)

        # Expected Delivery Date
        select_expected_delivery_date(page,7)

        # Delivery Location
        page.get_by_role("button", name="Select delivery location").click()
        page.get_by_role("option", name=DELIVERY_LOCATION_OPTION).click()

        page.get_by_role("button", name="Next").click()

        # ── Page 2: Item Details ────────────────────────────────────────────

        page.get_by_role("button").filter(has_text=re.compile(r"^$")).nth(2).click()
        page.get_by_role("textbox", name="Search items...").fill(row["Item Name"])
        page.get_by_role("option", name=row["Item Name"]).first.click()

        hsn_number = page.get_by_role("textbox", name="HSN Code is read-only").input_value()

        page.get_by_role("textbox", name="Qty").fill(str(row["Qty"]))

        page.get_by_role("button", name="Payment Terms", exact=True).click()
        page.get_by_role("textbox", name="Search...").fill(row["Payment Terms"])
        page.get_by_role("option", name=row["Payment Terms"]).click()

        # Negotiable toggle
        page.get_by_role("switch", name="Negotiable").click()

        page.get_by_role("button", name="Next").click()

        # ── Page 3: Vendor Selection ────────────────────────────────────────

        # Deselect the HSN chip
        page.get_by_role("button", name=str(hsn_number)).click()

        # Vendor
        page.get_by_role("textbox", name="Search vendors...").fill(row["vendor"])
        page.get_by_text(row["vendor"]).click()

        page.get_by_role("button", name="Next").click()

        # ── Page 4: Publish ─────────────────────────────────────────────────
        page.wait_for_timeout(1_000)

        # Screenshot before publish
        ss_path = take_screenshot(page, f"row_{row_index:03d}_{row['RFQ Title']}")
        result["Screenshot"] = ss_path

        page.get_by_role("button", name="Publish RFQ").nth(1).click()
        page.wait_for_load_state("networkidle", timeout=20_000)

        print(f"  ✅  Row {row_index}: {row['RFQ Title']} — PUBLISHED")

    except Exception as exc:
        result["Status"] = "FAIL"
        result["Error"]  = str(exc)
        # Still grab a screenshot on failure for debugging
        try:
            ss_path = take_screenshot(page, f"FAIL_row_{row_index:03d}_{row['RFQ Title']}")
            result["Screenshot"] = ss_path
        except Exception:
            pass
        print(f"  ❌  Row {row_index}: {row['RFQ Title']} — FAILED: {exc}")

    return result


# ─── EXCEL REPORT ─────────────────────────────────────────────────────────────

def generate_report(results: list[dict]):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        print("openpyxl not installed — run: pip install openpyxl")
        return

    wb = Workbook()

    # ── Summary Sheet ───────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    total  = len(results)
    passed = sum(1 for r in results if r["Status"] == "PASS")
    failed = total - passed

    header_fill  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    pass_fill    = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
    fail_fill    = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
    alt_fill     = PatternFill("solid", start_color="EEF2FF", end_color="EEF2FF")
    white_fill   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    title_fill   = PatternFill("solid", start_color="2E4FA3", end_color="2E4FA3")

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title block
    ws_sum.merge_cells("A1:N1")
    ws_sum["A1"] = "RFQ Automation — Test Execution Report"
    ws_sum["A1"].font      = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    ws_sum["A1"].fill      = title_fill
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 36

    ws_sum.merge_cells("A2:N2")
    ws_sum["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   Total: {total}   |   Passed: {passed}   |   Failed: {failed}"
    ws_sum["A2"].font      = Font(name="Arial", size=10, italic=True, color="FFFFFF")
    ws_sum["A2"].fill      = PatternFill("solid", start_color="3A5BB8", end_color="3A5BB8")
    ws_sum["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[2].height = 20

    # Column headers
    headers = [
        "#", "RFQ Title", "RFx Type", "Description",
        "Procurement Type", "Delivery Terms", "Department", "Cost Center",
        "Item Name", "Qty", "Payment Terms", "Vendor",
        "Status", "Error / Notes",
    ]
    ws_sum.append([])          # row 3 spacer
    ws_sum.append(headers)     # row 4

    for col_idx, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=4, column=col_idx)
        cell.value     = h
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws_sum.row_dimensions[4].height = 30

    # Data rows
    for i, r in enumerate(results, 1):
        row_num = 4 + i
        row_data = [
            r["Row"],
            r["RFQ Title"],
            r["RFx Type"],
            r["Description"],
            r["Procurement Type"],
            r["Delivery Terms"],
            r["Department"],
            r["Cost Center"],
            r["Item Name"],
            r["Qty"],
            r["Payment Terms"],
            r["Vendor"],
            r["Status"],
            r["Error"],
        ]
        ws_sum.append(row_data)
        row_fill = pass_fill if r["Status"] == "PASS" else fail_fill
        bg_fill  = alt_fill  if i % 2 == 0 else white_fill

        for col_idx in range(1, len(headers) + 1):
            cell = ws_sum.cell(row=row_num, column=col_idx)
            cell.font      = Font(name="Arial", size=9)
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx == 13:          # Status column
                cell.fill = row_fill
                cell.font = Font(name="Arial", size=9, bold=True,
                                 color="375623" if r["Status"] == "PASS" else "9C0006")
            else:
                cell.fill = bg_fill

    # Column widths
    widths = [5, 38, 20, 32, 17, 22, 16, 16, 28, 6, 32, 28, 8, 40]
    for col_idx, w in enumerate(widths, 1):
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = w

    ws_sum.freeze_panes = "A5"

    # ── Detail Sheet (one row per result, with embedded screenshot) ─────────
    ws_det = wb.create_sheet("Detailed Results")

    det_headers = ["#", "RFQ Title", "Status", "Timestamp", "Error", "Screenshot"]
    ws_det.append(det_headers)
    for col_idx, h in enumerate(det_headers, 1):
        cell = ws_det.cell(row=1, column=col_idx)
        cell.value     = h
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws_det.row_dimensions[1].height = 28

    img_row_height = 120   # pixels → approx 90 pt
    for i, r in enumerate(results, 1):
        row_num = i + 1
        ws_det.cell(row=row_num, column=1).value = r["Row"]
        ws_det.cell(row=row_num, column=2).value = r["RFQ Title"]
        ws_det.cell(row=row_num, column=3).value = r["Status"]
        ws_det.cell(row=row_num, column=4).value = r["Timestamp"]
        ws_det.cell(row=row_num, column=5).value = r["Error"]

        row_fill = pass_fill if r["Status"] == "PASS" else fail_fill
        for col_idx in range(1, 6):
            cell = ws_det.cell(row=row_num, column=col_idx)
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font      = Font(name="Arial", size=9)
            if col_idx == 3:
                cell.fill = row_fill
                cell.font = Font(name="Arial", size=9, bold=True,
                                 color="375623" if r["Status"] == "PASS" else "9C0006")

        # Embed screenshot thumbnail
        ss = r.get("Screenshot", "")
        if ss and os.path.exists(ss):
            try:
                img = XLImage(ss)
                img.width  = 200
                img.height = 112
                cell_addr  = f"F{row_num}"
                ws_det.add_image(img, cell_addr)
                ws_det.row_dimensions[row_num].height = 85
            except Exception:
                ws_det.cell(row=row_num, column=6).value = ss
        else:
            ws_det.cell(row=row_num, column=6).value = "No screenshot"

    ws_det.column_dimensions["A"].width = 5
    ws_det.column_dimensions["B"].width = 38
    ws_det.column_dimensions["C"].width = 10
    ws_det.column_dimensions["D"].width = 20
    ws_det.column_dimensions["E"].width = 45
    ws_det.column_dimensions["F"].width = 32
    ws_det.freeze_panes = "A2"

    # ── Stats Sheet ──────────────────────────────────────────────────────────
    ws_stats = wb.create_sheet("Stats")
    ws_stats.merge_cells("A1:C1")
    ws_stats["A1"] = "Execution Statistics"
    ws_stats["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws_stats["A1"].fill      = title_fill
    ws_stats["A1"].alignment = Alignment(horizontal="center")

    stats = [
        ("Total RFQs Processed", total),
        ("Passed",               passed),
        ("Failed",               failed),
        ("Pass Rate",            f"{passed/total*100:.1f}%" if total else "0%"),
    ]
    for idx, (label, val) in enumerate(stats, 3):
        ws_stats[f"A{idx}"] = label
        ws_stats[f"B{idx}"] = val
        ws_stats[f"A{idx}"].font = Font(name="Arial", bold=True, size=10)
        ws_stats[f"B{idx}"].font = Font(name="Arial", size=10)
        bg = pass_fill if "Pass" in label and label != "Pass Rate" else (
             fail_fill if label == "Failed" else white_fill)
        ws_stats[f"A{idx}"].fill = bg
        ws_stats[f"B{idx}"].fill = bg
    ws_stats.column_dimensions["A"].width = 28
    ws_stats.column_dimensions["B"].width = 18

    wb.save(REPORT_FILE)
    print(f"\n📊  Report saved → {REPORT_FILE}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    rows = load_csv(CSV_FILE)
    if not rows:
        print("❌  CSV is empty or missing.")
        sys.exit(1)

    print(f"📋  Loaded {len(rows)} RFQ rows from {CSV_FILE}")

    results: list[dict] = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=500)
        page    = browser.new_page()

        # ── Single login ───────────────────────────────────────────────────
        print("\n🔐  Logging in …")
        login(page)

        # Navigate to dashboard after login
        page.get_by_role("button").first.click()

        #go to RFQ Management 
        page.get_by_role("link", name="RFQ Management").click()

        # ── N iterations ───────────────────────────────────────────────────
        for idx, row in enumerate(rows, 1):
            print(f"\n▶  Processing row {idx}/{len(rows)}: {row.get('RFQ Title', '')}")
            result = submit_rfq(page, row, idx)
            results.append(result)
            time.sleep(1)  # brief pause between submissions

        browser.close()

    # ── Generate report ────────────────────────────────────────────────────
    print("\n📝  Generating Excel test report …")
    generate_report(results)
    print("✅  Done!")


if __name__ == "__main__":
    main()