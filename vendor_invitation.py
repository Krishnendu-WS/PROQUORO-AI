"""
Phase: Vendor Invitation
URL: stg.proquro.ai/company-admin/invite

Flow:
  Signin (auto) → navigate to dashboard
  → sidebar: click "Vendor Management" to expand
  → click "Invite Vendor" → lands on Vendor Invitations page
  → click "+ New Invitation"
  → fill form:
      Supplier Company Name  (LLM synthetic)
      Contact Person         (LLM synthetic)
      Phone Number           (LLM synthetic — 10-digit Indian)
      Email Address          (from vendor_emails.xlsx — first row where Invitation Sent is blank)
  → click "Send Invitation"
  → mark that email as sent (write "Yes" to Invitation Sent column)
  → verify success
"""

import asyncio
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from playwright.async_api import TimeoutError as PlaywrightTimeout

import config
from core.base_phase import BasePhase
from core.browser import BrowserManager
from core.data_manager import DataManager
from core.llm import LLMClient

INVITE_URL        = f"{config.BASE_URL}/company-admin/invite"
VENDOR_EMAILS_FILE = config.DATA_DIR / "vendor_emails.xlsx"


class VendorInvitationPhase(BasePhase):
    PHASE      = "vendor_invitation"
    TARGET_URL = INVITE_URL

    def __init__(self, company: str, browser: BrowserManager):
        super().__init__(company, browser)
        self.data_mgr = DataManager(company)
        self.llm      = LLMClient()

    # ── Entry point ───────────────────────────────────────────────────────────

    async def run_phase(self):
        await self._navigate_to_invite()
        await self._send_invitation()

    # ── Navigation ────────────────────────────────────────────────────────────

    async def _navigate_to_invite(self):
        step = "Navigate"

        # Try direct URL first
        await self.browser.navigate(INVITE_URL)
        await asyncio.sleep(2)

        if "invite" in self.page.url and "company-admin" in self.page.url:
            ss = await self.browser.screenshot(self.company, self.PHASE, "01_invite_page")
            self.reporter.log(step, "Vendor Invitations page loaded (direct URL)",
                              "PASS", self.page.url, ss)
            return

        # Sidebar fallback
        self.reporter.log(step, "Direct URL failed — trying sidebar", "INFO")
        await self._via_sidebar()

    async def _via_sidebar(self):
        step = "Navigate"
        await self.browser.navigate(config.DASHBOARD_URL)
        await asyncio.sleep(2)

        # Open sidebar hamburger
        opened = False
        for sel in [
            "button[aria-label*='menu']",
            "button[aria-label*='sidebar']",
            "[class*='hamburger']",
        ]:
            try:
                btn = await self.page.wait_for_selector(sel, timeout=2000, state="visible")
                if btn:
                    await btn.click()
                    await asyncio.sleep(1)
                    opened = True
                    break
            except PlaywrightTimeout:
                continue

        if not opened:
            await self.page.mouse.click(35, 85)
            await asyncio.sleep(1)

        # Scroll sidebar to VENDORS section
        await self.page.evaluate("""
            () => {
                const links = document.querySelectorAll('a, span, li, div');
                for (const el of links) {
                    if (el.textContent.trim() === 'Vendor Management') {
                        el.scrollIntoView({ behavior: 'smooth', block: 'center' });
                        break;
                    }
                }
            }
        """)
        await asyncio.sleep(0.5)

        # Click "Vendor Management" to expand it
        expanded = False
        for sel in [
            "text=Vendor Management",
            "a:has-text('Vendor Management')",
            "[href*='vendor']",
        ]:
            try:
                el = await self.page.wait_for_selector(sel, timeout=5000, state="visible")
                if el:
                    await el.click()
                    await asyncio.sleep(1)
                    expanded = True
                    self.reporter.log(step, "Vendor Management expanded", "INFO")
                    break
            except PlaywrightTimeout:
                continue

        if not expanded:
            self.reporter.log(step, "Could not expand Vendor Management", "FAIL")
            return

        ss = await self.browser.screenshot(self.company, self.PHASE, "01b_vendor_expanded")
        self.reporter.log(step, "Vendor Management section expanded", "INFO", "", ss)

        # Click "Invite Vendor" sub-item
        for sel in [
            "text=Invite Vendor",
            "a:has-text('Invite Vendor')",
            "[href*='invite']",
        ]:
            try:
                el = await self.page.wait_for_selector(sel, timeout=5000, state="visible")
                if el:
                    await el.click()
                    await self.page.wait_for_url("**/invite**", timeout=10000)
                    ss = await self.browser.screenshot(
                        self.company, self.PHASE, "01c_invite_page"
                    )
                    self.reporter.log(step, "Vendor Invitations via sidebar",
                                      "PASS", self.page.url, ss)
                    return
            except PlaywrightTimeout:
                continue

        self.reporter.log(step, "Could not navigate to Invite Vendor", "FAIL")

    # ── Send Invitation ───────────────────────────────────────────────────────

    async def _send_invitation(self):
        step = "SendInvitation"

        # Get LLM data (company name, contact, phone)
        data    = self._get_or_generate()
        company = data.get("supplier_company_name", "")
        contact = data.get("contact_person", "")
        phone   = data.get("phone_number", "")

        # Get next unused vendor email
        vendor_email, email_row = self._get_next_vendor_email()
        if not vendor_email:
            self.reporter.log(step, "No unused vendor emails available",
                              "FAIL", f"Check {VENDOR_EMAILS_FILE}")
            return

        self.reporter.log(step, "Vendor email selected", "INFO", vendor_email)

        # Click "+ New Invitation"
        clicked = False
        for sel in [
            "button:has-text('New Invitation')",
            "button:has-text('+ New Invitation')",
            "a:has-text('New Invitation')",
        ]:
            try:
                btn = await self.page.wait_for_selector(sel, timeout=5000, state="visible")
                if btn:
                    await btn.click()
                    self.reporter.log(step, "Clicked '+ New Invitation'", "PASS", sel)
                    clicked = True
                    break
            except PlaywrightTimeout:
                continue

        if not clicked:
            self.reporter.log(step, "Could not find '+ New Invitation' button", "FAIL")
            return

        await asyncio.sleep(1.2)
        ss = await self.browser.screenshot(self.company, self.PHASE, "02_form_open")
        self.reporter.log(step, "New Vendor Invitation form opened", "INFO", "", ss)

        # ── Fill fields ───────────────────────────────────────────────────────

        # Supplier Company Name
        await self._fill_field(
            "input[placeholder*='Supplier company name'], "
            "input[placeholder*='company name'], "
            "input[placeholder*='Company name']",
            company, step, "Supplier Company Name"
        )

        # Contact Person
        await self._fill_field(
            "input[placeholder*='Contact name'], "
            "input[placeholder*='contact person'], "
            "input[placeholder*='Contact Person']",
            contact, step, "Contact Person"
        )

        # Phone Number
        await self._fill_field(
            "input[placeholder*='9876543210'], "
            "input[placeholder*='Phone'], "
            "input[type='tel'], "
            "input[placeholder*='phone']",
            phone, step, "Phone Number"
        )

        # Email Address — use JS to find the input inside the modal specifically,
        # scoped under the "Email Address" label to avoid matching the background
        # search bar which also has an email-like input visible behind the modal.
        await self._fill_modal_email(vendor_email, step)

        ss = await self.browser.screenshot(self.company, self.PHASE, "03_form_filled")
        self.reporter.log(step, "Form filled", "INFO",
                          f"{company} | {contact} | {vendor_email}", ss)

        # ── Submit ────────────────────────────────────────────────────────────
        submitted = False
        for sel in [
            "button:has-text('Send Invitation')",
            "button:has-text('Send invitation')",
            "button[type='submit']:has-text('Send')",
        ]:
            try:
                btn = await self.page.wait_for_selector(sel, timeout=5000, state="visible")
                if btn:
                    await btn.click()
                    self.reporter.log(step, "Clicked 'Send Invitation'", "PASS", sel)
                    submitted = True
                    break
            except PlaywrightTimeout:
                continue

        if not submitted:
            self.reporter.log(step, "Could not click 'Send Invitation'", "FAIL")
            return

        await asyncio.sleep(2.5)
        ss = await self.browser.screenshot(self.company, self.PHASE, "04_after_submit")

        # ── Verify & mark email as sent ───────────────────────────────────────
        modal_gone = await self._modal_closed("New Vendor Invitation")
        toast_ok   = (
            await self._text_in_page("success")
            or await self._text_in_page("sent")
            or await self._text_in_page("Invitation sent")
        )

        if modal_gone or toast_ok:
            self._mark_email_sent(email_row)
            self.reporter.log(step, "Vendor invitation sent", "PASS",
                              f"Invitation sent to {vendor_email}", ss)
        else:
            error_msg = await self._get_validation_error()
            self.reporter.log(step, "Invitation may have failed", "FAIL",
                              error_msg or "Modal still open — check screenshot", ss)

    # ── Vendor email file helpers ─────────────────────────────────────────────

    def _get_next_vendor_email(self) -> tuple[str, int]:
        """
        Read vendor_emails.xlsx, return (email, row_index) for the first row
        where 'Invitation Sent' is blank/empty/None.
        Returns ('', -1) if none available or file missing.
        """
        if not VENDOR_EMAILS_FILE.exists():
            print(f"⚠️  vendor_emails.xlsx not found: {VENDOR_EMAILS_FILE}")
            print(f"   Please create it with columns: Email | Invitation Sent")
            return "", -1

        try:
            wb = openpyxl.load_workbook(str(VENDOR_EMAILS_FILE))
            ws = wb.active

            # Find column indices from header row
            headers = [
                str(c.value).strip().lower() if c.value else ""
                for c in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            try:
                email_col = next(
                    i for i, h in enumerate(headers) if "email" in h
                )
                sent_col = next(
                    i for i, h in enumerate(headers)
                    if "invitation" in h or "sent" in h
                )
            except StopIteration:
                print("⚠️  vendor_emails.xlsx missing expected columns (Email, Invitation Sent)")
                return "", -1

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                email = str(row[email_col]).strip() if row[email_col] else ""
                sent  = str(row[sent_col]).strip().lower() if row[sent_col] else ""
                if email and sent not in ("yes", "y", "true", "1", "sent"):
                    return email, row_idx

            print("⚠️  All vendor emails have already been used (Invitation Sent = Yes)")
            return "", -1

        except Exception as e:
            print(f"⚠️  Error reading vendor_emails.xlsx: {e}")
            return "", -1

    def _mark_email_sent(self, row_idx: int):
        """
        Write 'Yes' to the 'Invitation Sent' column for the given row.
        """
        if row_idx < 2 or not VENDOR_EMAILS_FILE.exists():
            return
        try:
            wb = openpyxl.load_workbook(str(VENDOR_EMAILS_FILE))
            ws = wb.active

            headers = [
                str(c.value).strip().lower() if c.value else ""
                for c in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            sent_col_idx = next(
                (i + 1 for i, h in enumerate(headers)
                 if "invitation" in h or "sent" in h),
                None
            )
            if sent_col_idx:
                ws.cell(row=row_idx, column=sent_col_idx).value = "Yes"
                wb.save(str(VENDOR_EMAILS_FILE))
                print(f"✅ Marked row {row_idx} as sent in vendor_emails.xlsx")
        except Exception as e:
            print(f"⚠️  Could not update vendor_emails.xlsx: {e}")

    # ── LLM data ──────────────────────────────────────────────────────────────

    def _get_or_generate(self) -> dict:
        key = "vendor_invitation"
        if self.data_mgr.has_step_data(key):
            data = self.data_mgr.get_step_data(key)
            self.reporter.log(key, "Data from cache", "INFO", f"{len(data)} fields")
            return data
        self.reporter.log(key, "Generating via LLM", "INFO")
        data = self.llm.generate_step_data(self.company, key)
        if data:
            self.data_mgr.save_step_data(key, data)
            self.reporter.log(key, "LLM data cached", "INFO", f"{len(data)} fields")
        return data

    # ── Field helpers ─────────────────────────────────────────────────────────

    async def _fill_modal_email(self, email: str, step: str):
        """
        Fill the Email Address field scoped strictly inside the modal dialog.
        Uses JS to find the input nearest to the 'Email Address' label inside
        the modal — avoids matching the background vendor search bar.
        """
        if not email:
            self.reporter.log_field(step, "Email Address", "", "SKIP", "No email")
            return
        try:
            filled = await self.page.evaluate(f"""
                () => {{
                    // Find the modal container first
                    const modal = document.querySelector(
                        '[role="dialog"], [class*="modal"], [class*="Dialog"], ' +
                        '[class*="sheet"], [class*="Sheet"]'
                    );
                    const root = modal || document;

                    // Find label with text "Email Address" inside modal
                    const labels = root.querySelectorAll('label, span, p, div');
                    for (const lbl of labels) {{
                        const t = lbl.textContent.trim().toLowerCase();
                        if (t === 'email address' || t.startsWith('email address')) {{
                            // Walk up to find the input sibling or child
                            let container = lbl.parentElement;
                            for (let i = 0; i < 4; i++) {{
                                if (!container) break;
                                const inp = container.querySelector(
                                    'input[type="email"], input[placeholder*="vendor"], ' +
                                    'input[placeholder*="company.com"]'
                                );
                                if (inp && inp.offsetParent !== null) {{
                                    inp.focus();
                                    inp.click();
                                    // Use React-compatible value setter
                                    const nativeInput = Object.getOwnPropertyDescriptor(
                                        window.HTMLInputElement.prototype, 'value'
                                    );
                                    nativeInput.set.call(inp, '{email}');
                                    inp.dispatchEvent(new Event('input', {{bubbles: true}}));
                                    inp.dispatchEvent(new Event('change', {{bubbles: true}}));
                                    return true;
                                }}
                                container = container.parentElement;
                            }}
                        }}
                    }}

                    // Fallback: find all email inputs inside modal, pick the one
                    // whose placeholder matches vendor@company.com exactly
                    const inputs = root.querySelectorAll('input');
                    for (const inp of inputs) {{
                        const ph = (inp.placeholder || '').toLowerCase();
                        if (ph.includes('vendor') || ph.includes('company.com')) {{
                            if (inp.offsetParent !== null) {{
                                inp.focus();
                                inp.click();
                                const nativeInput = Object.getOwnPropertyDescriptor(
                                    window.HTMLInputElement.prototype, 'value'
                                );
                                nativeInput.set.call(inp, '{email}');
                                inp.dispatchEvent(new Event('input', {{bubbles: true}}));
                                inp.dispatchEvent(new Event('change', {{bubbles: true}}));
                                return true;
                            }}
                        }}
                    }}
                    return false;
                }}
            """)
            if filled:
                self.reporter.log_field(step, "Email Address", email)
            else:
                self.reporter.log_field(step, "Email Address", "", "FAIL",
                                        "Could not find email input inside modal")
        except Exception as e:
            self.reporter.log_field(step, "Email Address", "", "FAIL", str(e))

    async def _fill_field(self, selector: str, value: str, step: str, name: str):
        if not value and value != 0:
            self.reporter.log_field(step, name, "", "SKIP", "No value")
            return
        value = str(value)  # always cast — LLM sometimes returns numbers
        try:
            el = await self.page.wait_for_selector(selector, timeout=6000, state="visible")
            if el:
                await el.evaluate("el => { el.focus(); el.click(); }")
                await asyncio.sleep(0.2)
                await el.fill(value)
                self.reporter.log_field(step, name, value)
        except Exception as e:
            self.reporter.log_field(step, name, "", "FAIL", str(e))

    async def _get_validation_error(self) -> str:
        """Scrape any visible validation error text from the modal."""
        try:
            errors = await self.page.evaluate("""
                () => {
                    const msgs = document.querySelectorAll(
                        '[class*="error"], [class*="invalid"], [class*="validation"], ' +
                        'p[style*="red"], span[style*="red"]'
                    );
                    return Array.from(msgs)
                        .filter(el => el.offsetParent !== null)
                        .map(el => el.textContent.trim())
                        .filter(t => t.length > 0)
                        .join(' | ');
                }
            """)
            return errors or ""
        except Exception:
            return ""

    async def _modal_closed(self, modal_text: str) -> bool:
        try:
            await self.page.wait_for_selector(
                f"text={modal_text}", state="hidden", timeout=5000
            )
            return True
        except Exception:
            return False

    async def _text_in_page(self, text: str) -> bool:
        if not text:
            return False
        try:
            await self.page.wait_for_selector(f"text={text}", timeout=3000)
            return True
        except Exception:
            return False